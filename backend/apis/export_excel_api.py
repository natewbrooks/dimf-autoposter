from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, Query
from fastapi.responses import FileResponse
import os
import tempfile
import subprocess
from datetime import datetime
from typing import Optional
from sqlalchemy.orm import Session

from database import get_db

router = APIRouter()

# Temporary directory for storing generated Excel files
TEMP_DIR = os.path.join(tempfile.gettempdir(), "dimf_exports")
os.makedirs(TEMP_DIR, exist_ok=True)

@router.get("/")
async def export_excel(
    background_tasks: BackgroundTasks,
    filename: Optional[str] = Query(None, description="Custom filename (without extension)"),
    db: Session = Depends(get_db)
):
    """
    Generate an Excel export of all database tables using ExportExcelService.
    Returns a downloadable Excel file.
    """
    # Create a unique filename
    if not filename:
        filename = "dimf-posts"
    
    # Add timestamp to prevent overwriting
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(TEMP_DIR, f"{filename}_{timestamp}.xlsx")
    
    try:
        # Generate the Excel file
        export_to_excel(db, filepath)
        
        # Schedule file cleanup
        background_tasks.add_task(cleanup_export_file, filepath, delay=3600)  # 1 hour
        
        # Set download filename (without the timestamp in the user-visible name)
        download_filename = f"{filename}.xlsx"
        
        return FileResponse(
            filepath, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=download_filename
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

def cleanup_export_file(filepath: str, delay: int = 3600):
    """Clean up temporary export files after a delay"""
    import time
    import os
    
    time.sleep(delay)
    if os.path.exists(filepath):
        try:
            os.remove(filepath)
        except:
            pass

def export_to_excel(db: Session, filepath: str):
    """
    Export all database tables to an Excel file
    """
    # Create a new workbook
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from sqlalchemy import text
    
    workbook = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    
    # Export each table
    export_posts(db, workbook)
    # export_images(db, workbook)
    export_post_images(db, workbook)
    export_post_distributions(db, workbook)
    export_platforms(db, workbook)
    # export_users(db, workbook)
    
    # Save the workbook
    workbook.save(filepath)
    
def create_header_row(sheet, headers):
    """Create a header row with bold formatting and background color"""
    from openpyxl.styles import Font, Alignment, PatternFill
    
    header_row = sheet.append(headers)
    
    # Style the header row
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
def export_posts(db, workbook):
    """Export Posts table data"""
    from sqlalchemy import text
    
    sheet = workbook.create_sheet("Posts")
    
    # Create header row
    headers = ["Post ID", "Name", "Date of Death", "Content", "Created By", "Created At"]
    create_header_row(sheet, headers)
    
    # Fetch data
    query = text("""
        SELECT p.PostID, p.Name, p.DateOfDeath, p.Content, p.CreatedBy, 
               u.Username as CreatedByUser, p.CreatedAt 
        FROM Posts p
        LEFT JOIN Users u ON p.CreatedBy = u.UserID
    """)
    results = db.execute(query).fetchall()
    
    # Fill data
    for row in results:
        created_by = f"{row.CreatedBy} ({row.CreatedByUser})" if row.CreatedByUser else row.CreatedBy
        sheet.append([
            row.PostID,
            row.Name,
            row.DateOfDeath,
            row.Content if row.Content else "",
            created_by,
            str(row.CreatedAt) if row.CreatedAt else ""
        ])
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)  # Cap at 100 to avoid extremely wide columns
        sheet.column_dimensions[column_letter].width = adjusted_width

def export_platforms(db, workbook):
    """Export SocialMediaPlatforms table data"""
    from sqlalchemy import text
    
    sheet = workbook.create_sheet("Platforms")
    
    # Create header row
    headers = ["Platform ID", "Name", "API Access Status", "Platform URL", "Icon URL"]
    create_header_row(sheet, headers)
    
    # Fetch data
    query = text("""
        SELECT PlatformID, Name, APIAccessStatus, PlatformURL, IconURL 
        FROM SocialMediaPlatforms
    """)
    results = db.execute(query).fetchall()
    
    # Fill data
    for row in results:
        sheet.append([
            row.PlatformID,
            row.Name,
            "Yes" if row.APIAccessStatus else "No",
            row.PlatformURL if row.PlatformURL else "",
            row.IconURL if row.IconURL else ""
        ])
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        sheet.column_dimensions[column_letter].width = adjusted_width

def export_images(db, workbook):
    """Export Images table data"""
    from sqlalchemy import text
    
    sheet = workbook.create_sheet("Images")
    
    # Create header row
    headers = ["Image ID", "URL", "Source"]
    create_header_row(sheet, headers)
    
    # Fetch data
    query = text("""
        SELECT ImageID, URL, Source 
        FROM Images
    """)
    results = db.execute(query).fetchall()
    
    # Fill data
    for row in results:
        sheet.append([
            row.ImageID,
            row.URL,
            row.Source if row.Source else ""
        ])
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        sheet.column_dimensions[column_letter].width = adjusted_width

def export_post_images(db, workbook):
    """Export PostImages join table data with related information"""
    from sqlalchemy import text
    
    sheet = workbook.create_sheet("Post Images")
    
    # Create header row
    headers = ["Post ID", "Image ID", "Post Name", "Image URL"]
    create_header_row(sheet, headers)
    
    # Fetch data with joins
    query = text("""
        SELECT pi.PostID, pi.ImageID, p.Name as PostName, i.URL as ImageURL 
        FROM PostImages pi 
        JOIN Posts p ON pi.PostID = p.PostID 
        JOIN Images i ON pi.ImageID = i.ImageID
    """)
    results = db.execute(query).fetchall()
    
    # Fill data
    for row in results:
        sheet.append([
            row.PostID,
            row.ImageID,
            row.PostName,
            row.ImageURL
        ])
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        sheet.column_dimensions[column_letter].width = adjusted_width

def export_post_distributions(db, workbook):
    """Export PostDistributions join table data with related information"""
    from sqlalchemy import text
    
    sheet = workbook.create_sheet("Post Distributions")
    
    # Create header row
    headers = ["Post ID", "Platform ID", "Post Name", "Platform Name"]
    create_header_row(sheet, headers)
    
    # Fetch data with joins
    query = text("""
        SELECT pd.PostID, pd.PlatformID, p.Name as PostName, smp.Name as PlatformName 
        FROM PostDistributions pd 
        JOIN Posts p ON pd.PostID = p.PostID 
        JOIN SocialMediaPlatforms smp ON pd.PlatformID = smp.PlatformID
    """)
    results = db.execute(query).fetchall()
    
    # Fill data
    for row in results:
        sheet.append([
            row.PostID,
            row.PlatformID,
            row.PostName,
            row.PlatformName
        ])
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        sheet.column_dimensions[column_letter].width = adjusted_width

def export_users(db, workbook):
    """Export Users table data (excluding passwords)"""
    from sqlalchemy import text
    
    sheet = workbook.create_sheet("Users")
    
    # Create header row
    headers = ["User ID", "Username", "Email"]
    create_header_row(sheet, headers)
    
    # Fetch data
    query = text("""
        SELECT UserID, Username, Email 
        FROM Users
    """)
    results = db.execute(query).fetchall()
    
    # Fill data
    for row in results:
        sheet.append([
            row.UserID,
            row.Username,
            row.Email if row.Email else ""
        ])
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        sheet.column_dimensions[column_letter].width = adjusted_width